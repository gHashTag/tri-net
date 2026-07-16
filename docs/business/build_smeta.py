from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

F_TITLE = Font(name='Arial', size=14, bold=True)
F_SUB   = Font(name='Arial', size=9, italic=True, color='555555')
F_HDR   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_SEC   = Font(name='Arial', size=10, bold=True)
F_CELL  = Font(name='Arial', size=10)
F_INPUT = Font(name='Arial', size=10, color='0000FF')
F_TOTAL = Font(name='Arial', size=11, bold=True)
F_NOTE  = Font(name='Arial', size=9, color='333333')

FILL_HDR   = PatternFill('solid', fgColor='1F3864')
FILL_SEC   = PatternFill('solid', fgColor='D9E1F2')
FILL_SUB   = PatternFill('solid', fgColor='FCE4D6')
FILL_TOTAL = PatternFill('solid', fgColor='F8CBAD')
FILL_KP    = PatternFill('solid', fgColor='FFF2CC')

THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RUB = '#,##0" ₽"'
INT = '#,##0'
AL = Alignment(horizontal='left', vertical='top', wrap_text=True)
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AR = Alignment(horizontal='right', vertical='center')

HDR = ["№","Наименование","Назначение / комментарий","Кол-во","Ед.","Цена ориент., ₽","Сумма ориент., ₽","Цена факт/КП, ₽","Сумма факт/КП, ₽","Поставщик / источник"]
WIDTHS = {'A':4,'B':46,'C':54,'D':7,'E':6,'F':16,'G':17,'H':16,'I':17,'J':30}

def setup(ws):
    for col,w in WIDTHS.items():
        ws.column_dimensions[col].width = w

def title_block(ws, lines, start=1):
    r = start
    for i, line in enumerate(lines):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(row=r, column=1, value=line)
        c.font = F_TITLE if i == 0 else F_SUB
        c.alignment = Alignment(wrap_text=True, vertical='center')
        r += 1
    return r + 1  # blank line after

def header_row(ws, r):
    for i, h in enumerate(HDR, start=1):
        c = ws.cell(row=r, column=i, value=h); c.font = F_HDR; c.fill = FILL_HDR; c.alignment = AC; c.border = BORDER
    return r + 1

def section_row(ws, r, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(row=r, column=1, value=text); c.font = F_SEC; c.fill = FILL_SEC
    for col in range(1, 11):
        ws.cell(row=r, column=col).border = BORDER
    return r + 1

def item_row(ws, r, n, name, comment, qty, unit, price, supplier):
    vals = [n, name, comment, qty, unit, price, f"=D{r}*F{r}", None, f"=D{r}*H{r}", supplier]
    for col, v in enumerate(vals, start=1):
        ws.cell(row=r, column=col, value=v)
    for col in range(1, 11):
        c = ws.cell(row=r, column=col); c.border = BORDER; c.font = F_CELL
        if col in (2, 3, 10): c.alignment = AL
        elif col in (1, 4, 5): c.alignment = AC
        else: c.alignment = AR
        if col in (6, 7, 8, 9): c.number_format = RUB
        if col == 4: c.number_format = INT
        if col in (4, 6): c.font = F_INPUT
        if col == 8: c.fill = FILL_KP; c.font = F_INPUT
    return r + 1

def subtotal_row(ws, r, label, first, last):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value=label); c.font = F_SEC; c.alignment = AR
    ws.cell(row=r, column=7, value=f"=SUM(G{first}:G{last})")
    ws.cell(row=r, column=9, value=f"=SUM(I{first}:I{last})")
    for col in range(1, 11):
        c = ws.cell(row=r, column=col); c.fill = FILL_SUB; c.border = BORDER
        if col in (7, 9): c.number_format = RUB; c.alignment = AR; c.font = F_SEC
    return r + 1

def grand_row(ws, r, label, gcells, icells):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value=label); c.font = F_TOTAL; c.alignment = AR
    ws.cell(row=r, column=7, value="=" + "+".join(gcells))
    ws.cell(row=r, column=9, value="=" + "+".join(icells))
    for col in range(1, 11):
        c = ws.cell(row=r, column=col); c.fill = FILL_TOTAL; c.border = BORDER
        if col in (7, 9): c.number_format = RUB; c.alignment = AR; c.font = F_TOTAL
    return r + 1

def write_sections(ws, r, sections):
    """sections: list of dict(title, optional, items). Returns (next_row, list of (subtotal_g_cell, optional))."""
    subs = []
    n = 1
    for sec in sections:
        r = section_row(ws, r, sec['title'])
        first = r
        for it in sec['items']:
            r = item_row(ws, r, n, *it); n += 1
        last = r - 1
        sub_r = r
        r = subtotal_row(ws, r, sec['subtotal'], first, last)
        subs.append((f"G{sub_r}", f"I{sub_r}", sec.get('optional', False)))
    return r, subs

def notes_block(ws, r, notes):
    for i, t in enumerate(notes):
        c = ws.cell(row=r, column=1, value=t)
        c.font = Font(name='Arial', size=10, bold=True) if i == 0 else F_NOTE
        c.alignment = Alignment(horizontal='left', vertical='top')
        r += 1
    return r

# ---------------- Sheet 1: MVP ----------------
wb = Workbook()
ws = wb.active
ws.title = "Смета MVP"
setup(ws)
r = title_block(ws, [
    "Смета на оборудование — MVP (минимальная mesh-сеть, 3 узла)",
    "Проект: Trinity Secure Mesh Node v2.0 / TRI-1  ·  DOI 10.5281/zenodo.19227877",
    "Конфигурация: 3× FPGA-узла ALINX AX7203 — 3-узловой mesh + видео/аудио-тракт для демо «видео-рации» (UC-1 / AC-2)",
    "Валюта: RUB. «Ориент.» — рыночные оценки (требуют КП). Жёлтый столбец «факт/КП» — для цен поставщиков; суммы пересчитываются.",
])
r = header_row(ws, r)
sections_mvp = [
 {'title': "1. ОСНОВНОЕ ОБОРУДОВАНИЕ (FPGA-узлы)",
  'subtotal': "Подытог 1, ₽",
  'items': [
   ("Плата FPGA ALINX AX7203 (Xilinx Artix-7 XC7A200T)", "Узел mesh: packet-fabric, 2× GbE, HDMI in/out, 4× GTP, audit-log в DDR3. Базовая плата всех 3 узлов", 3, "шт", 55000, "ALINX RU-партнёр (alinx.com)"),
  ]},
 {'title': "2. ВИДЕО/АУДИО-ТРАКТ (демо видео-рации — UC-1 / AC-2)",
  'subtotal': "Подытог 2, ₽",
  'items': [
   ("Камера-модуль OV5640 (ALINX) или HDMI-камера 1080p", "Источник видео (HDMI in / expansion-порт). 2 оконечных узла видео-рации; 3-й узел — ретранслятор", 2, "шт", 5000, "дистрибьютор / ALINX"),
   ("HDMI-монитор (можно существующие)", "Отображение принятого видеопотока (HDMI out)", 2, "шт", 12000, "локальный / существующий"),
   ("HDMI-кабель 1–2 м", "Подключение камеры/монитора к HDMI in/out (по 2 на узел)", 4, "шт", 600, "локальный дистрибьютор"),
   ("Микрофон + динамик / USB-гарнитура", "Голосовой канал H.264/AAC, приоритет QoS. По 1 комплекту на оконечный узел", 2, "компл", 2500, "локальный дистрибьютор"),
  ]},
 {'title': "3. СЕТЬ И АКСЕССУАРЫ",
  'subtotal': "Подытог 3, ₽",
  'items': [
   ("Карта microSD 32 ГБ (industrial)", "Загрузка битстрима / хранение audit-логов на узле", 3, "шт", 1200, "локальный дистрибьютор"),
   ("Патч-корд Ethernet Cat6, 1–3 м", "Mesh-линки между узлами (триангуляция) + host/запас", 6, "шт", 400, "локальный дистрибьютор"),
   ("USB-JTAG загрузчик (резерв)", "Прошивка FPGA. На AX7203 обычно есть онбордный USB-JTAG — 1 шт. про запас", 1, "шт", 6000, "ALINX / дистрибьютор"),
   ("Блок питания 12 В", "Входит в комплект платы AX7203", 3, "шт", 0, "в комплекте"),
   ("Кабель USB-UART", "Консоль/отладка. Входит в комплект платы", 3, "шт", 0, "в комплекте"),
  ]},
 {'title': "4. ОПЦИОНАЛЬНО / ПРИ НЕОБХОДИМОСТИ",
  'optional': True,
  'subtotal': "Подытог 4 (опц.), ₽",
  'items': [
   ("SFP оптический модуль 1G", "Для GTP/HSST оптических линков. На AX7203 опц. (mesh идёт по GbE); обязателен для Pango-порта", 2, "шт", 2500, "локальный дистрибьютор"),
   ("Рабочая станция (host PC)", "Проверка корня Merkle (audit-port), мониторинг mesh. Обычно существующая", 1, "шт", 120000, "при отсутствии"),
   ("Коммутатор GbE управляемый, 8 портов", "Агрегация host/audit-порта при звезде мониторинга (опц.)", 1, "шт", 9000, "локальный дистрибьютор"),
   ("Модуль КМ211 K1986ВО1Т", "Хранитель ключа на host-стороне (крипто-граница, R-6). Опц. для MVP", 1, "шт", 4000, "www2.km211.ru"),
  ]},
]
r, subs = write_sections(ws, r, sections_mvp)
req_g  = [g for g, i, opt in subs if not opt]
req_i  = [i for g, i, opt in subs if not opt]
opt_g  = [g for g, i, opt in subs if opt]
opt_i  = [i for g, i, opt in subs if opt]
r = grand_row(ws, r, "ИТОГО (рабочий MVP, разд. 1–3), ₽", req_g, req_i)
r = grand_row(ws, r, "ИТОГО (опционально, разд. 4), ₽", opt_g, opt_i)
r = grand_row(ws, r, "ВСЕГО, ₽", req_g + opt_g, req_i + opt_i)
r += 1
notes_block(ws, r, [
 "Примечания:",
 "1. MVP = 3 FPGA-узла ALINX AX7203 (Artix-7 XC7A200T). 3 узла — минимум для mesh (многохоповая маршрутизация A→B→C); 2 узла — только point-to-point.",
 "2. Раздел 2 (видео/аудио) ОБЯЗАТЕЛЕН: критерий AC-2 требует передачи видео/аудио с приоритетом голоса. Продукт — «видео-рация» (UC-1), без камеры/монитора/микрофона демо не собирается.",
 "3. Цены «ориент.» — рыночные оценки, подлежат уточнению. Импортные платы (ALINX) дистрибьютор котирует в USD/CNY с конверсией в ₽.",
 "4. Host-CPU (Эльбрус / Baikal / ПК) не включён как обязательная закупка — узел садится рядом с существующим хостом по PCIe / USB-UART.",
 "5. Полный перечень остального железа из документа (Pango P100/P390, ВЗПП-С, CPE510, LoRa, БПЛА, камеры, тапаут) — на листе «Расширение (вне MVP)».",
 "6. Источник номенклатуры: КТП «Trinity Secure Mesh Node v2.0», разделы 5.1, 6, 7, 8.",
 "7. Покрываемые критерии приёмки: AC-1 (битстрим), AC-2 (mesh + видео/аудио), AC-3 (audit-log/anti-replay), AC-4 (16 опкодов), AC-5 (GF(16)), AC-8 (TRI_SEAL/G_MERKLE).",
])
ws.freeze_panes = 'A7'

# ---------------- Sheet 2: Extension (out of MVP) ----------------
ws2 = wb.create_sheet("Расширение (вне MVP)")
setup(ws2)
r = title_block(ws2, [
    "Оборудование следующих этапов (вне MVP) — справочно",
    "Весь остальный перечень железа из КТП (разд. 5.1, 6, 7, 9, 10). Цены — ГРУБЫЕ оценки, обязателен запрос КП.",
    "Цель листа — полнота (ничего не потеряно), а не точный бюджет. Стоимость Pango / ВЗПП-С / тапаута сильно варьируется.",
])
r = header_row(ws2, r)
sections_ext = [
 {'title': "Ветка A — RU-порт и серия (Ф5 / серийный узел)",
  'subtotal': "Подытог (ветка A), ₽",
  'items': [
   ("Плата PangoMicro P100 / AXP100 (Logos-2, 99 900 LUT4)", "RU-friendly пилот (AC-6): 1 GB DDR3, PCIe 2.0, 8× HSST 6.6 Gbps, 2× SFP + GbE", 1, "шт", 180000, "MacroGroup (macrogroup.ru) — запрос КП"),
   ("Плата PangoMicro P390 / AXP390 (Titan-2 PG2T390H, 365 400 LUT4)", "Серийный кии-узел: 8 GB DDR4, PCIe 3.0, 16× HSST 13.125 Gbps, 276 I/O, SFP + GbE", 1, "шт", 350000, "fpgapro.com — запрос КП"),
   ("ВЗПП-С / КТЦ 5578ТС064 (55 856 LE, аналог Altera EP3C55)", "Compliance / урезанный packet-guard: 2.3 Mbit BRAM, 156 умножителей, GMII soft", 1, "шт", 120000, "ВЗПП-С / eandc.ru — запрос КП"),
  ]},
 {'title': "Ветка B — кремний TRI-1 (silicon, услуги фаундри)",
  'subtotal': "Подытог (ветка B), ₽",
  'items': [
   ("Слот тапаута TinyTapeout TTSKY26b (SKY130)", "Дедлайн 17.05.2026; fallback TTSKY26c (авг. 2026). Возврат кристаллов 16.12.2026", 1, "слот", 90000, "app.tinytapeout.com/shuttles — тариф"),
   ("MPW-подача IHP SG13G2 (параллельно)", "Открытый PDK SG13G2; страховка от срыва тапаута (R-3)", 1, "подача", 250000, "IHP MPW — тариф (уточнять)"),
  ]},
 {'title': "Полевой mesh + БПЛА (Ф6, UC-1…UC-4)",
  'subtotal': "Подытог (поле/БПЛА), ₽",
  'items': [
   ("Плата ALINX AX7203 — доп. узлы (до mesh 8 узлов)", "Расширение с 3 до 8 узлов (финальное демо: 8 mesh + 1 TRI-1)", 5, "шт", 55000, "ALINX RU-партнёр"),
   ("CPE510 (5 ГГц outdoor-мост) + антенна", "Резервный полевой радиоканал (R-11), погодоустойчивость", 4, "шт", 6000, "локальный дистрибьютор"),
   ("LoRa-модуль + антенна", "Control-plane / fallback низкоскоростной командный канал против РЭБ (R-12)", 4, "шт", 2500, "локальный дистрибьютор"),
   ("БПЛА (носитель узла)", "Рой 4–16 бортов (UC-1). Минимум для полевого теста — 4. Стоимость сильно варьируется", 4, "шт", 80000, "подбор под поле"),
   ("Камера для БПЛА (aerial recon)", "Источник видеопотока на борту (Figure 6 «THE EYE»)", 4, "шт", 8000, "локальный дистрибьютор"),
   ("Корпус узла (enclosure, 2× GbE + JTAG)", "Полевой корпус узла (Figure 5 «THE NODE»), тепловой режим", 8, "шт", 5000, "локальный / под заказ"),
  ]},
]
r, subs2 = write_sections(ws2, r, sections_ext)
allg = [g for g, i, opt in subs2]
alli = [i for g, i, opt in subs2]
r = grand_row(ws2, r, "ИТОГО (справочно, грубо), ₽", allg, alli)
r += 1
notes_block(ws2, r, [
 "Примечания:",
 "• Этот лист — НЕ часть MVP. Он перечисляет всё остальное оборудование из документа для следующих этапов.",
 "• Цены Pango (P100/P390), ВЗПП-С и услуг тапаута/MPW — грубые оценки; обязателен запрос КП / тарифа фаундри.",
 "• Стоимость БПЛА и полевой обвязки зависит от выбранной платформы и требований поля.",
])
ws2.freeze_panes = 'A6'

wb.save("/Users/ssdm4/Desktop/PROJECTS/CLAUDE/Smeta_MVP_Trinity_Mesh_3node.xlsx")
print("saved")
